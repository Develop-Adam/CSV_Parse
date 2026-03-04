#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import load_workbook


#========================================================+=========================================#
'''
Helpers
'''
def resolve_path(script_dir: Path, p: str) -> Path:
    """
    Resolve a path relative to the script directory unless it is already absolute.
    """
    path = Path(p)
    return path if path.is_absolute() else (script_dir / path).resolve()


def str_bool(v: str) -> bool:
    """
    Convert common spreadsheet boolean strings to python bool.
    Accepts: TRUE/FALSE, true/false, 1/0, yes/no
    """
    if v is None:
        return False
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "on"):
        return True
    if s in ("false", "0", "no", "n", "off"):
        return False
    raise ValueError(f"Invalid boolean value: {v!r}")


def parse_skip_rows(v) -> List[int]:
    """
    Parse skip_rows from Excel value.
    Supports:
      - "0" or "0,1,2"
      - 0 (numeric)
      - empty -> []
    """
    if v is None:
        return []
    if isinstance(v, (int, float)) and str(v).strip() != "":
        return [int(v)]
    s = str(v).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(",") if p.strip() != ""]
    out = []
    for p in parts:
        out.append(int(p))
    return out


def read_kv_table(ws, start_row: int, key_col: int = 1, val_col: int = 2) -> Dict[str, Any]:
    """
    Read a key/value table starting at start_row where:
      row start_row contains headers: key | value
      subsequent rows contain key/value until a blank key is encountered.
    """
    # Expect header row
    h1 = ws.cell(row=start_row, column=key_col).value
    h2 = ws.cell(row=start_row, column=val_col).value
    if str(h1).strip().lower() != "key" or str(h2).strip().lower() != "value":
        raise ValueError(f"Expected 'key'/'value' header at row {start_row}, got {h1!r}/{h2!r}")

    out: Dict[str, Any] = {}
    r = start_row + 1
    while True:
        k = ws.cell(row=r, column=key_col).value
        v = ws.cell(row=r, column=val_col).value
        if k is None or str(k).strip() == "":
            break
        out[str(k).strip()] = v
        r += 1
    return out


def find_row_with_value(ws, value: str, col: int = 1, max_rows: int = 500) -> int:
    """
    Find the row index where ws.cell(row, col) equals 'value' (case-insensitive).
    Returns row index or raises if not found.
    """
    target = value.strip().lower()
    for r in range(1, max_rows + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        if str(v).strip().lower() == target:
            return r
    raise ValueError(f"Could not find '{value}' in column {col} of sheet '{ws.title}'")


def read_header_list(ws, start_row: int, col: int = 1) -> List[str]:
    """
    Read header list where:
      ws.cell(start_row, col) == 'header'
      values continue downward until a blank cell
    """
    h = ws.cell(row=start_row, column=col).value
    if str(h).strip().lower() != "header":
        raise ValueError(f"Expected 'header' at row {start_row}, got {h!r}")
    headers: List[str] = []
    r = start_row + 1
    while True:
        v = ws.cell(row=r, column=col).value
        if v is None or str(v).strip() == "":
            break
        headers.append(str(v).strip())
        r += 1
    return headers


#========================================================+=========================================#
'''
Excel -> settings.json generator

Workbook format:
- Sheet "Profiles": columns [profile_name, is_default]
- One sheet per profile name:
  - A section titled "CSV Settings" followed by key/value table
  - A section titled "Reporting Column Mapping" followed by key/value table
  - A section titled "Header List" followed by a single "header" column list
'''
def build_settings_from_excel(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)

    if "Profiles" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'Profiles' sheet")

    ws_profiles = wb["Profiles"]

    # Read profiles table
    # Expect header row: profile_name | is_default
    h1 = ws_profiles.cell(row=1, column=1).value
    h2 = ws_profiles.cell(row=1, column=2).value
    if str(h1).strip().lower() != "profile_name" or str(h2).strip().lower() != "is_default":
        raise ValueError("Profiles sheet must have headers: profile_name, is_default in row 1")

    profiles: List[str] = []
    default_profile: str | None = None

    r = 2
    while True:
        name = ws_profiles.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            break
        name = str(name).strip()
        profiles.append(name)

        is_def = ws_profiles.cell(row=r, column=2).value
        if is_def is not None and str(is_def).strip() != "":
            if str_bool(is_def):
                if default_profile and default_profile != name:
                    raise ValueError(f"Multiple default profiles set: {default_profile} and {name}")
                default_profile = name
        r += 1

    if not profiles:
        raise ValueError("No profiles found in Profiles sheet")

    if not default_profile:
        # Fall back to first profile if none marked default
        default_profile = profiles[0]

    settings: Dict[str, Any] = {
        "default_profile": default_profile,
        "profiles": {}
    }

    # For each profile sheet, parse sections
    for prof in profiles:
        if prof not in wb.sheetnames:
            raise ValueError(f"Missing profile sheet named '{prof}' (must match Profiles.profile_name)")

        ws = wb[prof]

        # Find section titles in column A
        csv_title_row = find_row_with_value(ws, "CSV Settings", col=1)
        rep_title_row = find_row_with_value(ws, "Reporting Column Mapping", col=1)
        hdr_title_row = find_row_with_value(ws, "Header List", col=1)

        # Key/value tables start on the row immediately after the section title
        csv_kv = read_kv_table(ws, start_row=csv_title_row + 1)
        rep_kv = read_kv_table(ws, start_row=rep_title_row + 1)

        # Header list starts on the row immediately after "Header List" title,
        # and expects the 'header' label in column A
        headers = read_header_list(ws, start_row=hdr_title_row + 1, col=1)

        # Normalize CSV kv into the expected JSON shape
        csv_json: Dict[str, Any] = {}

        if "delimiter" in csv_kv and csv_kv["delimiter"] is not None:
            csv_json["delimiter"] = str(csv_kv["delimiter"])

        if "encoding" in csv_kv and csv_kv["encoding"] is not None:
            csv_json["encoding"] = str(csv_kv["encoding"])

        if "skip_rows" in csv_kv:
            csv_json["skip_rows"] = parse_skip_rows(csv_kv.get("skip_rows"))

        if "header_from_settings" in csv_kv:
            csv_json["header_from_settings"] = str_bool(csv_kv.get("header_from_settings"))
        else:
            csv_json["header_from_settings"] = True  # sensible default

        # Always store header list
        csv_json["header"] = headers

        # Reporting kv: store as strings where present
        reporting_json: Dict[str, Any] = {}
        for k, v in rep_kv.items():
            if v is None:
                continue
            s = str(v).strip()
            if s == "":
                continue
            reporting_json[str(k).strip()] = s

        settings["profiles"][prof] = {
            "csv": csv_json,
            "reporting": reporting_json,
        }

    return settings


#========================================================+=========================================#
'''
CLI
'''
def build_parser():
    p = argparse.ArgumentParser(description="Convert Excel settings workbook into settings.json (profiles format).")
    p.add_argument("--excel", default="inputs/settings_template.xlsx",
                   help="Input Excel workbook (default: inputs/settings_template.xlsx)")
    p.add_argument("--output", default="inputs/settings.json",
                   help="Output settings.json path (default: inputs/settings.json)")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON with indentation")
    return p


#========================================================+=========================================#
'''
Main
'''
def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    script_dir = Path(__file__).resolve().parent
    excel_path = resolve_path(script_dir, args.excel)
    output_path = resolve_path(script_dir, args.output)

    if not excel_path.exists():
        print(f"[ERROR] Excel file not found: {excel_path}", file=sys.stderr)
        return 2

    try:
        settings = build_settings_from_excel(excel_path)
    except Exception as e:
        print(f"[ERROR] Failed to build settings from Excel: {e}", file=sys.stderr)
        return 3

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with output_path.open("w", encoding="utf-8") as f:
            if args.pretty:
                json.dump(settings, f, indent=2, ensure_ascii=False)
            else:
                json.dump(settings, f, ensure_ascii=False)
    except Exception as e:
        print(f"[ERROR] Failed to write JSON: {e}", file=sys.stderr)
        return 4

    print(f"[INFO] Read Excel: {excel_path}")
    print(f"[INFO] Wrote settings.json: {output_path}")
    print(f"[INFO] default_profile: {settings.get('default_profile')}")
    print(f"[INFO] profiles: {list((settings.get('profiles') or {}).keys())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
